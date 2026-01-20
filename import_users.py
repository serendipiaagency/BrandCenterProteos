import openpyxl
import sqlite3
import secrets
import string
import json
from datetime import datetime

# Mapping de columnas de productos a IDs de marcas
BRAND_MAPPING = {
    'Pack \nHA 1.5': 2,  # pbserum
    'Pack \nHA 2.0': 2,
    'Solutions \nHA 2.0': 2,
    'HA \nCORRECTOR': 2,
    'WAID': 3,
    'SHS30+\nHIGH': 2,
    'SPECIFIC': 2,
    'PLUS': 2,
    'SMARTKER': 2,
    'USA \nSPECIFIC': 2,
    'USA \nSMARTKER': 2
}

def generate_password(length=12):
    """Genera una contraseña aleatoria segura"""
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    password = ''.join(secrets.choice(alphabet) for i in range(length))
    return password

def process_excel_to_users():
    # Leer Excel
    wb = openpyxl.load_workbook('/home/user/BRAND CENTER - TABLA PERMISOS.xlsx')
    ws = wb.active
    
    # Obtener headers
    headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
    
    users_data = []
    
    # Procesar cada fila
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value:
                row_data[header] = str(cell_value).strip()
        
        if not row_data.get('MAIL DISTRIBUIDOR'):
            continue
        
        # Extraer datos del usuario
        name = row_data.get('NOMBRE', '')
        email = row_data.get('MAIL DISTRIBUIDOR', '').split(',')[0].strip()  # Primer email si hay múltiples
        market = row_data.get('MERCADO', '')
        country = row_data.get('PAÍS', '')
        distributor = row_data.get('DISTRIBUIDOR', '')
        
        if not email or not name:
            continue
        
        # Generar contraseña
        password = generate_password()
        
        # Determinar marcas con acceso
        brands_access = []
        for brand_col, brand_id in BRAND_MAPPING.items():
            if row_data.get(brand_col) == '√':
                if brand_id not in brands_access:
                    brands_access.append(brand_id)
        
        user = {
            'name': name,
            'email': email,
            'password': password,
            'role': 'distributor',
            'region': market,
            'country': country,
            'distributor': distributor,
            'language': 'ING',  # Inglés por defecto
            'brands_access': brands_access
        }
        
        users_data.append(user)
    
    return users_data

def insert_users_to_db(users_data):
    """Inserta usuarios en la base de datos SQLite local"""
    conn = sqlite3.connect('/home/user/webapp/.wrangler/state/v3/d1/miniflare-D1DatabaseObject/fc50b649db51ed0c303ff2c4b7c0eca2da269cc3dfc7ce40615fc37a7b53366c.sqlite')
    cursor = conn.cursor()
    
    inserted = 0
    skipped = 0
    
    for user in users_data:
        try:
            # Verificar si el usuario ya existe
            cursor.execute('SELECT id FROM users WHERE email = ?', (user['email'],))
            existing = cursor.fetchone()
            
            if existing:
                # Actualizar distributor si existe
                cursor.execute('''
                    UPDATE users SET distributor = ? WHERE email = ?
                ''', (user['distributor'], user['email']))
                print(f"🔄 Actualizado: {user['email']} - Distribuidor: {user['distributor']}")
                skipped += 1
                continue
            
            # Insertar usuario
            cursor.execute('''
                INSERT INTO users (email, password_hash, name, role, region, country, language, brands_access, distributor, active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', (
                user['email'],
                user['password'],  # En producción usar bcrypt
                user['name'],
                user['role'],
                user['region'],
                user['country'],
                user['language'],
                json.dumps(user['brands_access']),
                user['distributor']
            ))
            
            print(f"✅ Creado: {user['name']} ({user['email']}) - Password: {user['password']}")
            inserted += 1
            
        except Exception as e:
            print(f"❌ Error con {user['email']}: {str(e)}")
    
    conn.commit()
    conn.close()
    
    return inserted, skipped

# Ejecutar
print("🔄 Procesando Excel...")
users = process_excel_to_users()
print(f"📊 Total usuarios encontrados: {len(users)}")
print()

print("💾 Insertando en base de datos...")
inserted, skipped = insert_users_to_db(users)
print()
print(f"✅ Usuarios creados: {inserted}")
print(f"⚠️  Usuarios omitidos (ya existentes): {skipped}")
print()
print("🔐 IMPORTANTE: Guarda las contraseñas mostradas arriba")
